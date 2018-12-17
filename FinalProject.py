#VIT Python 2018 Final Project
#group no - 81
#Kunal Kokate - 16103B0021
#Neeraj Khare - 16103B0013
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl import load_workbook
import random
import smtplib

wb=load_workbook('VITPython18-FinalProject-Workbook.xlsx')
sheet=wb['UserDetails']
i=0
j=0
g=0
h=0

while(i!=3 and j!=3):
    i=i+1
    x=input("Enter Full Name:")
    k=2
    for k in range(2,7):
        if x==sheet.cell(row=k,column=4).value:
            print("Full name found")
            h=k
            print(h)
            while j!=3:
                j+=1
                y=input("Enter your bank account no.:")
                if y==sheet.cell(row=k,column=3).value:
                    print("Account number found")
                    print("Stage 1 successful")
                    g=1
                    break
                else:
                    print("Invalid User")

    if g==1:
        break
    print("Invalid user")

    if j==3:
        sheet.cell(row=h,column=7).value='Blocked'

if i==3 or j==3:
    print("Quit program is stage 1")
wb.save('VITPython18-FinalProject-Workbook.xlsx')
#stage2
number=random.randint(1000,9999)
print(number)
server = smtplib.SMTP('smtp.gmail.com',587)
server.starttls()
server.login("kunalkokate2753@gmail.com", "Hidaddy181998")
msg=str(number)
server.sendmail("kunalkokate2753@gmail.com", "kunalkokate2753@gmail.com",msg)
z=0
while z<3:
    
    inumber=input("Please enter OTP:")

    if msg==inumber:
        print("Correct...proceed to stage 3")
        break
    else:
        print("Incorrect")
        sheet.cell(row=h,column=7).value='Blocked'
        wb.save('VITPython18-FinalProject-Workbook.xlsx')
    z=z+1

#stage 3
ask=str(input("You want to debit 'd' or credit 'c':"))


if ask=='c' or ask=='C':
    f=sheet.cell(row=h,column=6).value
    print("Current Balance:",f)
    credit=int(input("Enter amount:"))
    sheet.cell(row=h,column=6).value+=credit
    print("transaction of ",credit, "is successful")
    print("Balance after transaction is:",sheet.cell(row=h,column=6).value)
    wb.save('VITPython18-FinalProject-Workbook.xlsx')

        
elif ask=='d' or ask=='D':
    f=sheet.cell(row=h,column=6).value
    print("Current Balance:",f)
    debit=int(input("Enter amount:"))
    sheet.cell(row=h,column=6).value-=debit
    print("transaction of ",debit, "is successful")
    print("Balance after transaction is:",sheet.cell(row=h,column=6).value)
    wb.save('VITPython18-FinalProject-Workbook.xlsx')

else :
    print("Error")